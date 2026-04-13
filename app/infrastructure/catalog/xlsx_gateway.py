from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(slots=True)
class XlsxCatalogConfig:
    workbook_path: str
    sheet_name: str = "Catalog"


class XlsxCatalogGateway:
    def __init__(self, config: XlsxCatalogConfig) -> None:
        self._config = config

    async def fetch_rows(self) -> list[dict[str, str]]:
        workbook_path = Path(self._config.workbook_path)
        if not workbook_path.exists():
            raise FileNotFoundError(f"Catalog workbook was not found: {workbook_path}")

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            if self._config.sheet_name not in workbook.sheetnames:
                raise KeyError(f"Sheet '{self._config.sheet_name}' was not found in {workbook_path.name}")
            sheet = workbook[self._config.sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return []

            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            entries: list[dict[str, str]] = []
            for row in rows[1:]:
                payload = {
                    header: "" if idx >= len(row) or row[idx] is None else str(row[idx]).strip()
                    for idx, header in enumerate(headers)
                    if header
                }
                if any(value for value in payload.values()):
                    entries.append(payload)
            return entries
        finally:
            workbook.close()
