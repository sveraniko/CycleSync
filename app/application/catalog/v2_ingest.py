from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.application.catalog.ingest import CatalogIngestService
from app.application.catalog.normalization import normalize_text, parse_bool, parse_decimal
from app.application.catalog.schemas import (
    CatalogProductInput,
    IngredientInput,
    IngestIssue,
    MediaInput,
    SourceLinkInput,
)


@dataclass(slots=True)
class WorkbookV2Sheets:
    products: list[dict[str, str]]
    ingredients: list[dict[str, str]]
    sources: list[dict[str, str]]
    media: list[dict[str, str]]
    aliases: list[dict[str, str]]


def _read_sheet(workbook_path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Missing required V2 sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        payload: list[dict[str, str]] = []
        for row in rows[1:]:
            item = {
                header: "" if idx >= len(row) or row[idx] is None else str(row[idx]).strip()
                for idx, header in enumerate(headers)
                if header
            }
            if any(item.values()):
                payload.append(item)
        return payload
    finally:
        workbook.close()


def read_workbook_v2(workbook_path: str) -> WorkbookV2Sheets:
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"Catalog workbook was not found: {path}")
    return WorkbookV2Sheets(
        products=_read_sheet(path, "Products"),
        ingredients=_read_sheet(path, "Ingredients"),
        sources=_read_sheet(path, "Sources"),
        media=_read_sheet(path, "Media"),
        aliases=_read_sheet(path, "Aliases"),
    )


def _to_decimal(value: str | None) -> Decimal | None:
    return parse_decimal(value)


def _to_int(value: str | None) -> int:
    try:
        return int((value or "0").strip())
    except ValueError:
        return 0


def build_v2_inputs(sheets: WorkbookV2Sheets) -> tuple[list[CatalogProductInput], list[IngestIssue]]:
    issues: list[IngestIssue] = []
    ingredients_by_product: dict[str, list[dict[str, str]]] = {}
    aliases_by_product: dict[str, list[str]] = {}
    media_by_product: dict[str, list[dict[str, str]]] = {}
    sources_by_product: dict[str, list[dict[str, str]]] = {}

    for row in sheets.ingredients:
        key = normalize_text(row.get("product_key", ""))
        if key.lower().startswith("hint:"):
            continue
        if not key:
            issues.append(IngestIssue(row_key=str(row.get("ingredient_order") or "?"), message="Ingredients.product_key missing"))
            continue
        ingredients_by_product.setdefault(key, []).append(row)

    for row in sheets.aliases:
        key = normalize_text(row.get("product_key", ""))
        alias = normalize_text(row.get("alias", ""))
        if key and alias:
            aliases_by_product.setdefault(key, []).append(alias)

    for row in sheets.media:
        key = normalize_text(row.get("product_key", ""))
        if key:
            media_by_product.setdefault(key, []).append(row)

    for row in sheets.sources:
        key = normalize_text(row.get("product_key", ""))
        if key:
            sources_by_product.setdefault(key, []).append(row)

    products: list[CatalogProductInput] = []
    for row in sheets.products:
        product_key = normalize_text(row.get("product_key", ""))
        if product_key.lower().startswith("hint:"):
            continue
        if not product_key:
            issues.append(IngestIssue(row_key="?", message="Products.product_key missing"))
            continue
        release_form = normalize_text(row.get("release_form", "")) or None
        ingredients_rows = sorted(
            ingredients_by_product.get(product_key, []),
            key=lambda item: int(item.get("ingredient_order") or "0"),
        )
        if not ingredients_rows:
            issues.append(IngestIssue(row_key=product_key, message="Product has no ingredient rows"))
            continue

        mapped_ingredients: list[IngredientInput] = []
        for ingredient in ingredients_rows:
            basis = normalize_text(ingredient.get("basis", ""))
            half_life = _to_decimal(ingredient.get("half_life_days"))
            active_fraction = _to_decimal(ingredient.get("active_fraction"))
            per_ml = _to_decimal(ingredient.get("amount_per_ml_mg"))
            per_unit = _to_decimal(ingredient.get("amount_per_unit_mg"))

            if basis not in {"per_ml", "per_unit"}:
                issues.append(IngestIssue(row_key=product_key, message=f"Invalid ingredient basis: {basis or 'empty'}"))
                continue
            if half_life is None:
                issues.append(IngestIssue(row_key=product_key, message="Ingredient missing half_life_days"))
                continue
            if active_fraction is None:
                issues.append(IngestIssue(row_key=product_key, message="Ingredient missing active_fraction"))
                continue
            if basis == "per_ml" and per_ml is None:
                issues.append(IngestIssue(row_key=product_key, message="Ingredient basis=per_ml but amount_per_ml_mg missing"))
                continue
            if basis == "per_unit" and per_unit is None:
                issues.append(IngestIssue(row_key=product_key, message="Ingredient basis=per_unit but amount_per_unit_mg missing"))
                continue

            mapped_ingredients.append(
                IngredientInput(
                    ingredient_name=normalize_text(ingredient.get("ingredient_name", "")),
                    qualifier=None,
                    amount=per_ml if basis == "per_ml" else per_unit,
                    unit="mg",
                    basis=basis,
                    half_life_days=half_life,
                    dose_guidance_min_mg_week=_to_decimal(ingredient.get("dose_guidance_min_mg_week")),
                    dose_guidance_max_mg_week=_to_decimal(ingredient.get("dose_guidance_max_mg_week")),
                    dose_guidance_typical_mg_week=_to_decimal(ingredient.get("dose_guidance_typical_mg_week")),
                    is_pulse_driver=parse_bool(ingredient.get("is_pulse_driver")),
                    parent_substance=normalize_text(ingredient.get("parent_substance", "")) or None,
                    ester_name=normalize_text(ingredient.get("ester_name", "")) or None,
                    amount_per_ml_mg=per_ml,
                    amount_per_unit_mg=per_unit,
                    active_fraction=active_fraction,
                    tmax_hours=_to_decimal(ingredient.get("tmax_hours")),
                    release_model=normalize_text(ingredient.get("release_model", "")) or None,
                    pk_notes=normalize_text(ingredient.get("pk_notes", "")) or None,
                )
            )

        if not mapped_ingredients:
            continue

        is_injectable = (release_form or "").startswith("injectable")
        volume_per_package_ml = _to_decimal(row.get("volume_per_package_ml"))
        units_per_package = _to_decimal(row.get("units_per_package"))
        if is_injectable and volume_per_package_ml is None:
            issues.append(IngestIssue(row_key=product_key, message="Injectable product missing volume_per_package_ml"))
            continue
        if not is_injectable and release_form in {"tablet", "capsule"} and units_per_package is None:
            issues.append(IngestIssue(row_key=product_key, message="Solid product missing units_per_package"))
            continue

        media_rows = media_by_product.get(product_key, [])
        media_items = [
            MediaInput(
                media_kind=normalize_text(r.get("media_kind", "")) or "image",
                ref=normalize_text(r.get("ref", "")),
                priority=_to_int(r.get("priority")),
                is_cover=parse_bool(r.get("is_cover")) is True,
                source_layer=normalize_text(r.get("source_layer", "")) or "import",
                is_active=parse_bool(r.get("is_active")) is not False,
            )
            for r in media_rows
            if normalize_text(r.get("ref", ""))
        ]
        image_refs = [m.ref for m in media_items if m.media_kind == "image" and m.is_active]
        video_refs = [m.ref for m in media_items if m.media_kind in {"video", "gif", "animation"} and m.is_active]
        source_rows = sources_by_product.get(product_key, [])
        source_links = [
            SourceLinkInput(
                kind=normalize_text(r.get("source_kind", "")) or "source",
                label=normalize_text(r.get("label", "")) or "Source",
                url=normalize_text(r.get("url", "")),
                priority=_to_int(r.get("priority")),
                source_layer=normalize_text(r.get("source_layer", "")) or "import",
                is_active=parse_bool(r.get("is_active")) is not False,
            )
            for r in source_rows
            if normalize_text(r.get("url", ""))
        ]

        products.append(
            CatalogProductInput(
                source_row_key=product_key,
                product_key=product_key,
                brand_name=normalize_text(row.get("brand", "")),
                display_name=normalize_text(row.get("display_name", "")),
                trade_name=normalize_text(row.get("trade_name", "")),
                release_form=release_form,
                concentration_raw=None,
                concentration_value=None,
                concentration_unit="mg",
                concentration_basis="per_ml" if is_injectable else "per_unit",
                official_url=normalize_text(row.get("official_url", "")) or None,
                authenticity_notes=normalize_text(row.get("authenticity_notes", "")) or None,
                max_injection_volume_ml=None,
                is_automatable=True,
                pharmacology_notes=normalize_text(row.get("pharmacology_notes", "")) or None,
                composition_basis_notes=None,
                package_kind=normalize_text(row.get("package_kind", "")) or None,
                volume_per_package_ml=volume_per_package_ml,
                unit_strength_mg=None,
                units_per_package=units_per_package,
                source_payload=row,
                aliases=aliases_by_product.get(product_key, []),
                ingredients=mapped_ingredients,
                image_refs=image_refs,
                video_refs=video_refs,
                source_links=source_links,
                media_items=media_items,
            )
        )

    product_keys = {item.product_key for item in products if item.product_key}
    for ingredient in sheets.ingredients:
        key = normalize_text(ingredient.get("product_key", ""))
        if key.lower().startswith("hint:"):
            continue
        if key and key not in product_keys:
            issues.append(IngestIssue(row_key=key, message="Ingredient references unknown product_key"))
    for source in sheets.sources:
        key = normalize_text(source.get("product_key", ""))
        if key.lower().startswith("hint:"):
            continue
        if key and key not in product_keys:
            issues.append(IngestIssue(row_key=key, message="Source references unknown product_key"))
    for media in sheets.media:
        key = normalize_text(media.get("product_key", ""))
        if key.lower().startswith("hint:"):
            continue
        if key and key not in product_keys:
            issues.append(IngestIssue(row_key=key, message="Media references unknown product_key"))

    return products, issues


@dataclass(slots=True)
class CatalogIngestServiceV2(CatalogIngestService):
    workbook_path: str

    async def run(self):  # type: ignore[override]
        sheets = read_workbook_v2(self.workbook_path)
        products, issues = build_v2_inputs(sheets)
        if issues:
            messages = "; ".join(issue.message for issue in issues[:5])
            raise ValueError(f"V2 workbook validation failed: {messages}")
        return await CatalogIngestService.run_from_products(self, products)
